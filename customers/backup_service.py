import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Customer, Expense, ExpenseCategory, Payment, PaymentDue,
    PaymentReminder, Reminder, ReminderSchedule, ServicePlan, WhatsAppMessage,
)


def _collect_section(data, headers, rows, filter_row=None):
    """Append a named section (headers + rows) to the backup data dict."""
    if filter_row:
        rows = [r for r in rows if filter_row(r)]
    data.append({
        'title': headers[0] if headers else 'Data',
        'headers': headers,
        'rows': rows,
    })


def collect_backup_data():
    """Gather all backup-worthy data from the database. Customers come first."""
    data = []
    today = datetime.now()

    _collect_section(
        data,
        ['Customers', 'Username', 'First Name', 'Last Name', 'Join Date', 'Phone',
         'Area', 'Street', 'St#', 'House#', 'Modem', 'Plan', 'Install Date',
         'Status', 'Notes', 'Created At', 'Updated At'],
        [
            (c.username, c.first_name, c.last_name, c.join_date.isoformat() if c.join_date else '',
             c.phone, c.address_area or '', c.street_name or '', c.street_num or '',
             c.house_num or '', c.modem_type, c.service_plan.name if c.service_plan else '',
             c.service_installation_date.isoformat() if c.service_installation_date else '',
             c.status, c.notes or '',
             c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
             c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else '')
            for c in Customer.objects.select_related('service_plan').all()
        ],
    )

    _collect_section(
        data,
        ['Service Plans', 'Name', 'Speed', 'Price', 'Active'],
        [(p.name, p.speed, float(p.price), 'Yes' if p.is_active else 'No') for p in ServicePlan.objects.all()],
    )

    _collect_section(
        data,
        ['Payments', 'Invoice', 'Customer', 'Amount', 'Payment Date', 'Month For',
         'Method', 'Received By', 'Notes'],
        [
            (p.invoice_number, p.customer.username, float(p.amount),
             p.payment_date.isoformat(), p.month_for.strftime('%B %Y'),
             p.get_method_display(), p.received_by.username, p.notes or '')
            for p in Payment.objects.select_related('customer', 'received_by').all()
        ],
    )

    _collect_section(
        data,
        ['Expense Categories', 'Name'],
        [(c.name,) for c in ExpenseCategory.objects.all()],
    )

    _collect_section(
        data,
        ['Expenses', 'Category', 'Description', 'Amount', 'Date'],
        [
            (e.category.name, e.description or '', float(e.amount),
             e.date.isoformat())
            for e in Expense.objects.select_related('category').all()
        ],
    )

    _collect_section(
        data,
        ['Reminders', 'Customer', 'Due Date', 'Type', 'Scheduled', 'Status', 'Message'],
        [
            (r.customer.username, r.due_date.isoformat(), r.get_reminder_type_display(),
             r.scheduled_time.isoformat(), r.status, r.message or '')
            for r in Reminder.objects.select_related('customer').all()
        ],
    )

    _collect_section(
        data,
        ['Reminder Schedule', 'Customer', 'Due Date', 'Send Time', 'Message', 'Sent'],
        [
            (r.customer.username, r.due_date.isoformat(), r.send_time.isoformat(),
             r.message, 'Yes' if r.sent else 'No')
            for r in ReminderSchedule.objects.select_related('customer').all()
        ],
    )

    _collect_section(
        data,
        ['WhatsApp Messages', 'Customer', 'Message', 'Timestamp', 'Status'],
        [
            (m.customer.username, m.message, m.timestamp.isoformat(), m.status)
            for m in WhatsAppMessage.objects.select_related('customer').all()
        ],
    )

    _collect_section(
        data,
        ['Payment Dues', 'Customer', 'Due Date', 'Amount', 'Paid'],
        [
            (d.customer.username, d.due_date.isoformat(), float(d.amount),
             'Yes' if d.is_paid else 'No')
            for d in PaymentDue.objects.select_related('customer').all()
        ],
    )

    _collect_section(
        data,
        ['Payment Reminders', 'Customer', 'Due Date', 'Type', 'Sent At', 'Sent'],
        [
            (r.customer.username, r.due_date.isoformat(), r.get_reminder_type_display(),
             r.sent_at.isoformat() if r.sent_at else '', 'Yes' if r.is_sent else 'No')
            for r in PaymentReminder.objects.select_related('customer').all()
        ],
    )

    return data, today


def _excel_style(ws, header, ncols):
    ws.freeze_panes = 'A2'
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4E73DF', end_color='4E73DF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 20
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            12, min(40, len(str(header[col - 1])) + 4))


def generate_excel_backup():
    data, now = collect_backup_data()
    wb = Workbook()
    wb.remove(wb.active)
    stamp = now.strftime('%Y%m%d_%H%M%S')

    for section in data:
        headers = section['headers']
        ws = wb.create_sheet(title=section['title'][:31])
        ws.append(headers)
        for row in section['rows']:
            ws.append([str(x) if x is not None else '' for x in row])
        _excel_style(ws, headers, len(headers))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, f'backup_{stamp}.xlsx', now


def generate_pdf_backup():
    data, now = collect_backup_data()
    buffer = io.BytesIO()
    stamp = now.strftime('%Y%m%d_%H%M%S')

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f'SSISP Backup {stamp}',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'BackupTitle', parent=styles['Title'], fontSize=18, spaceAfter=4)
    subtitle_style = ParagraphStyle(
        'BackupSub', parent=styles['Normal'], fontSize=10,
        textColor=colors.grey, spaceAfter=14)
    heading_style = ParagraphStyle(
        'BackupHeading', parent=styles['Heading2'], fontSize=13,
        textColor=colors.HexColor('#4E73DF'), spaceBefore=12, spaceAfter=6)
    cell_style = ParagraphStyle(
        'BackupCell', fontSize=7, leading=8)
    header_cell_style = ParagraphStyle(
        'BackupHeaderCell', fontSize=7, leading=8,
        textColor=colors.white, fontName='Helvetica-Bold')

    elements = [Paragraph('SSISP Data Backup', title_style),
                Paragraph(
                    f'Generated on {now.strftime("%d %b %Y at %H:%M:%S")} — '
                    f'customer, billing and expense records', subtitle_style)]

    for section in data:
        elements.append(Paragraph(section['title'], heading_style))
        headers = [Paragraph(h, header_cell_style) for h in section['headers']]
        body = []
        for row in section['rows']:
            body.append([Paragraph(str(x) if x is not None else '', cell_style) for x in row])

        if not body:
            elements.append(Paragraph('No records.', cell_style))
            continue

        table = Table([headers] + body, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E73DF')),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D4D9')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F8')]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 8))

    doc.build(elements)
    buffer.seek(0)
    return buffer, f'backup_{stamp}.pdf', now